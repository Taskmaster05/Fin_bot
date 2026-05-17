import uuid
import re
import os
import tempfile
import operator
from typing import TypedDict, Annotated

from langgraph.graph import StateGraph, START, END
from langgraph.checkpoint.memory import InMemorySaver
from langchain_core.runnables import RunnableConfig
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from agent.tools import pinecone_search, stock_query, llm
from agent.prompts import (
    ROUTER_PROMPT,
    FORMAT_PROMPT,
    CONTEXT_RESOLVER_PROMPT,
    SYNTHESISER_PROMPT
)


# ── State ──────────────────────────────────────────────────

class AgentState(TypedDict):
    messages:      Annotated[list, operator.add]
    tool_results:  list
    sources:       Annotated[list, operator.add]
    output_format: str
    _route:        str
    download_path: str
    download_type: str


# ── Node 1: Context resolver ───────────────────────────────

def context_resolver_node(state: AgentState):
    messages = state["messages"]
    latest   = messages[-1]["content"]

    if len(messages) <= 1:
        return {}

    history_str = "\n".join(
        f"{m['role'].upper()}: {m['content'][:300]}"
        for m in messages[-7:-1]
    )

    resolved = llm.invoke(
        CONTEXT_RESOLVER_PROMPT.format(
            history=history_str,
            question=latest
        )
    ).content.strip()

    updated = messages[:-1] + [{
        "role": "user",
        "content": resolved,
        "_original": latest
    }]
    return {"messages": updated}


# ── Node 2: Router ─────────────────────────────────────────

def router_node(state: AgentState):
    question = state["messages"][-1]["content"]
    route    = llm.invoke(ROUTER_PROMPT.format(question=question)).content.strip().lower()
    if route not in ("stock", "docs", "both"):
        route = "docs"
    return {"_route": route}


# ── Node 3: Format classifier ──────────────────────────────

def format_classifier_node(state: AgentState):
    question = state["messages"][-1]["content"]
    fmt      = llm.invoke(FORMAT_PROMPT.format(question=question)).content.strip().lower()
    if fmt not in ("table", "excel", "pdf", "text"):
        fmt = "text"
    return {"output_format": fmt}


# ── Node 4: Tool executor ──────────────────────────────────

def tool_executor_node(state: AgentState):
    route   = state["_route"]
    results = []

    if route in ("docs", "both"):
        results.append(pinecone_search.invoke(
            {"question": state["messages"][-1]["content"]}
        ))

    if route in ("stock", "both"):
        results.append(stock_query.invoke(
            {"question": state["messages"][-1]["content"]}
        ))

    return {"tool_results": results}


# ── Node 5: Synthesiser ────────────────────────────────────

def synthesiser_node(state: AgentState):
    all_sources    = []
    context_blocks = []

    for item in state["tool_results"]:
        context_blocks.append(item.get("content", ""))
        all_sources.extend(item.get("sources", []))

    unique_sources = list(dict.fromkeys(all_sources))
    source_map     = {s: f"[{i+1}]" for i, s in enumerate(unique_sources)}
    source_map_str = "\n".join(f"{tag} = {src}" for src, tag in source_map.items())

    context  = "\n\n".join(context_blocks)
    question = state["messages"][-1]["content"]

    response = llm.invoke(
        SYNTHESISER_PROMPT.format(
            source_map=source_map_str,
            context=context,
            question=question
        )
    ).content

    legend = "\n\n---\n**Sources:**\n" + "\n".join(
        f"{tag} {src}" for src, tag in source_map.items()
    )
    final = response + legend

    return {
        "messages":     [{"role": "assistant", "content": final}],
        "sources":      unique_sources,
        "tool_results": []
    }


# ── Node 6: Output renderer ────────────────────────────────

def render_output_node(state: AgentState):
    fmt     = state["output_format"]
    content = state["messages"][-1]["content"]
    sources = state["sources"]

    if fmt == "excel":
        path = _render_excel(content, sources)
        return {"download_path": path, "download_type": "excel"}

    elif fmt == "pdf":
        path = _render_pdf(content, sources)
        return {"download_path": path, "download_type": "pdf"}

    return {}


def _render_excel(content: str, sources: list) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Data"

    lines = [l for l in content.split("\n") if "|" in l]
    row_idx = 1
    for line in lines:
        cells = [c.strip() for c in line.split("|") if c.strip()]
        if all(set(c) <= set("-: ") for c in cells):
            continue
        for col_idx, cell in enumerate(cells, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell)
        row_idx += 1

    if row_idx == 1:
        ws["A1"] = content

    ws2 = wb.create_sheet("Sources")
    ws2["A1"] = "Sources used"
    for i, src in enumerate(sources, 2):
        ws2.cell(row=i, column=1, value=src)

    tmpdir = tempfile.gettempdir()
    path = os.path.join(tmpdir, f"financial_{uuid.uuid4().hex[:8]}.xlsx")
    wb.save(path)
    return path


def _render_pdf(content: str, sources: list) -> str:
    tmpdir = tempfile.gettempdir()
    path = os.path.join(tmpdir, f"financial_{uuid.uuid4().hex[:8]}.pdf")
    doc = SimpleDocTemplate(path, pagesize=A4)
    styles = getSampleStyleSheet()
    story  = []

    for para in content.split("\n\n"):
        clean = re.sub(r"\*\*(.*?)\*\*", r"<b>\1</b>", para.strip())
        clean = re.sub(r"\*(.*?)\*",     r"<i>\1</i>", clean)
        if clean:
            story.append(Paragraph(clean, styles["Normal"]))
            story.append(Spacer(1, 10))

    story.append(Paragraph("Sources", styles["Heading2"]))
    for src in sources:
        story.append(Paragraph(f"• {src}", styles["Normal"]))

    doc.build(story)
    return path


# ── Build graph ────────────────────────────────────────────

def build_graph():
    g = StateGraph(AgentState)

    g.add_node("context_resolver",  context_resolver_node)
    g.add_node("router",            router_node)
    g.add_node("format_classifier", format_classifier_node)
    g.add_node("tool_executor",     tool_executor_node)
    g.add_node("synthesiser",       synthesiser_node)
    g.add_node("render_output",     render_output_node)

    g.add_edge(START,               "context_resolver")
    g.add_edge("context_resolver",  "router")
    g.add_edge("router",            "format_classifier")
    g.add_edge("format_classifier", "tool_executor")
    g.add_edge("tool_executor",     "synthesiser")
    g.add_edge("synthesiser",       "render_output")
    g.add_edge("render_output",     END)

    checkpointer = InMemorySaver()
    return g.compile(checkpointer=checkpointer)


agent = build_graph()